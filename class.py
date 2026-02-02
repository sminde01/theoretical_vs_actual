import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from collections import defaultdict
import io
import os
import hashlib


# =============================================
# GLOBAL CONFIGURATION
# =============================================
STEEL_BUY_RATE = 55.0
STEEL_SCRAP_RATE = 35.0
THEORETICAL_MARGIN = 0.1

def get_file_hash(uploaded_file):
    """Generate hash for uploaded file to use as cache key"""
    uploaded_file.seek(0)
    file_hash = hashlib.md5(uploaded_file.read()).hexdigest()
    uploaded_file.seek(0)
    return file_hash

# =============================================
# FLATTENER CLASS
# =============================================
class DataFlattener:
    """Handles flattening operations for different hierarchy levels"""
    
    def __init__(self, df):
        self.df = df.copy()
        self.df.fillna('', inplace=True)
        self.additional_columns = self._get_additional_columns()


    def _ensure_string_part_numbers(self, row_data):
        """Ensure part numbers are stored as strings"""
        for key in row_data.keys():
            if 'Part No' in key and row_data[key] != '-':
                row_data[key] = str(row_data[key])
        return row_data


    def detect_hierarchy_level(self):
        """Automatically detect the number of hierarchy levels"""
        if 'LEVEL' not in self.df.columns:
            raise ValueError("LEVEL column not found in the data")
        
        max_level = self.df['LEVEL'].max()
        
        filtered_df = self.df[
            ~self.df['Type'].astype(str).str.strip().str.lower().isin(['fastner', 'dummy part', 'bop'])
        ]
        
        max_valid_level = filtered_df['LEVEL'].max()
        
        if max_valid_level <= 1:
            return 2  # 2-level hierarchy (Assembly → Component)
        elif max_valid_level == 2:
            return 3  # 3-level: SA → Assembly → Component
        elif max_valid_level >= 3:
            return 4  # 4-level hierarchy (SA → Module → Assembly → Component)
        else:
            return 2  # Default to 2-level if unclear


    def _get_additional_columns(self):
        """Get ALL columns that should be preserved in flattened output"""
        exclude_cols = {'LEVEL', 'PART NO', 'PART NAME', 'Type', 'Validity', 'MOD'}
        return [col for col in self.df.columns if col not in exclude_cols]
    
    def flatten_2_levels(self):
        """Flatten with Assembly and Component levels (Level 0 and 1)"""
        flat_data = []
        assembly = None
        
        for _, row in self.df.iterrows():
            level = int(row['LEVEL']) if pd.notna(row['LEVEL']) else -1
            typ = str(row['Type']).strip().lower()
            part_no = str(row['PART NO'])
            
            if typ in ['fastner', 'dummy part', 'bop']:
                continue
            
            if level == 0:
                assembly = {
                    'Part No': str(part_no),
                    'Name': row['PART NAME'],
                    'Validity': row.get('Validity') if pd.notna(row.get('Validity')) else 1
                }
            elif level == 1:
                row_data = {
                    'Assembly Part No': str(assembly['Part No']) if assembly else '-',
                    'Assembly Name': assembly['Name'] if assembly else '-',
                    'Assembly Validity': assembly['Validity'] if assembly else 1,
                    'Component Part No': str(part_no),
                    'Component Name': row['PART NAME'],
                    'Component Validity': row.get('Validity') if pd.notna(row.get('Validity')) else 1,
                }
                
                # Add all other columns from the row
                for col in self.df.columns:
                    if col not in ['LEVEL', 'PART NO', 'PART NAME', 'Type', 'Validity', 'MOD']:
                        if col not in row_data:
                            row_data[col] = row.get(col, '')
                
                flat_data.append(row_data)
        
        return pd.DataFrame(flat_data)


    def flatten_3_levels(self):
        """Flatten with SA, Assembly, and Component levels (Level 0, 1, 2)"""
        flat_data = []
        visited_edges = set()
        sa = assembly = None
        
        for _, row in self.df.iterrows():
            level = int(row['LEVEL']) if pd.notna(row['LEVEL']) else -1
            typ = str(row['Type']).strip().lower() if pd.notna(row['Type']) else ''
            part_no = str(row['PART NO'])
            
            if typ in ['fastner', 'dummy part', 'bop']:
                continue
            
            if level == 0:
                sa = {
                    'Part No': str(part_no),
                    'Name': row['PART NAME'],
                    'Validity': row.get('Validity') if pd.notna(row.get('Validity')) else 1
                }
                assembly = None
            
            elif level == 1:
                assembly = {
                    'Part No': str(part_no),
                    'Name': row['PART NAME'],
                    'Validity': row.get('Validity') if pd.notna(row.get('Validity')) else 1
                }
            
            elif level == 2:
                # Component under assembly
                edge_key = (
                    sa['Part No'] if sa else '-',
                    assembly['Part No'] if assembly else '-',
                    part_no
                )
                if edge_key in visited_edges:
                    continue
                visited_edges.add(edge_key)
                
                row_data = {
                    'SA Part No': str(sa['Part No']) if sa else '-',
                    'SA Name': sa['Name'] if sa else '-',
                    'SA Validity': sa['Validity'] if sa else 1,
                    'Assembly Part No': str(assembly['Part No']) if assembly else '-',
                    'Assembly Name': assembly['Name'] if assembly else '-',
                    'Assembly Validity': assembly['Validity'] if assembly else 1,
                    'Component Part No': str(part_no),
                    'Component Name': row['PART NAME'],
                    'Component Validity': row.get('Validity') if pd.notna(row.get('Validity')) else 1,
                }
                
                for col in self.additional_columns:
                    row_data[col] = row.get(col, '')
                
                flat_data.append(row_data)
        
        return pd.DataFrame(flat_data)


    def flatten_4_levels(self):
        """Flatten with SA, Module, Assembly, and Component levels"""
        flat_data = []
        sa = module = assembly = None
        
        for _, row in self.df.iterrows():
            level = int(row['LEVEL']) if pd.notna(row['LEVEL']) else -1
            typ = str(row['Type']).strip().lower() if pd.notna(row['Type']) else ''
            part_no = str(row['PART NO'])
            
            if typ in ['fastner', 'dummy part', 'bop']:
                continue
            
            if level == 0:
                sa = {
                    'Part No': str(part_no),
                    'Name': row['PART NAME'],
                    'Validity': row.get('Validity') if pd.notna(row.get('Validity')) else 1
                }
                module = assembly = None
            elif level == 1:
                module = {
                    'Part No': str(part_no),
                    'Name': row['PART NAME'],
                    'Validity': row.get('Validity') if pd.notna(row.get('Validity')) else 1
                }
                assembly = None
            elif level == 2:
                if typ == 'assembly':
                    assembly = {
                        'Part No': str(part_no),
                        'Name': row['PART NAME'],
                        'Validity': row.get('Validity') if pd.notna(row.get('Validity')) else 1
                    }
                else:
                    row_data = self._create_4level_row_data(sa, module, None, row)
                    flat_data.append(row_data)
            elif level == 3:
                row_data = self._create_4level_row_data(sa, module, assembly, row)
                flat_data.append(row_data)
        
        return pd.DataFrame(flat_data)

    def flatten_auto(self):
        """Automatically detect hierarchy and flatten accordingly"""
        hierarchy_level = self.detect_hierarchy_level()
        
        if hierarchy_level == 4:
            return self.flatten_4_levels(), True  
        elif hierarchy_level == 3:
            return self.flatten_3_levels(), False 
        else:
            return self.flatten_2_levels(), False

    def _create_4level_row_data(self, sa, module, assembly, row):
        """Create row data for 4-level hierarchy"""
        row_data = {
            'SA Part No': str(sa['Part No']) if sa else '-',
            'SA Name': sa['Name'] if sa else '-',
            'SA Validity': sa['Validity'] if sa else 1,
            'Module Part No': str(module['Part No']) if module else '-',
            'Module Name': module['Name'] if module else '-',
            'Module Validity': module['Validity'] if module else 1,
            'Assembly Part No': assembly['Part No'] if assembly else '-',
            'Assembly Name': assembly['Name'] if assembly else '-',
            'Assembly Validity': assembly['Validity'] if assembly else 1,
            'Component Part No': str(row['PART NO']),
            'Component Name': row['PART NAME'],
            'Component Validity': row.get('Validity') if pd.notna(row.get('Validity')) else 1,
        }
        

        for col in self.df.columns:
            if col not in ['LEVEL', 'PART NO', 'PART NAME', 'Type', 'Validity', 'MOD']:
                if col not in row_data:  
                    row_data[col] = row.get(col, '')
        
        return row_data
    


# =============================================
# WEIGHT CALCULATOR CLASS
# =============================================
class WeightCalculator:
    """Handles all weight-related calculations"""
    
    @staticmethod
    def calculate_input_weight_theoretical(row):
        """Calculate theoretical input weight based on commodity type"""
        commodity = str(row['Commodity']).strip().lower()
        no_of_components = row['NO OF COMPONENTS']
        if pd.isna(no_of_components) or no_of_components == 0:
            return np.nan
        
        try:
            if commodity == 'plate':
                return (row['THK MM'] * row['SHEET WIDTH MM'] * row['SHEET LENGTH MM'] * 7.85 * 1e-6) / row['NO OF COMPONENTS']
            elif commodity in ['pipe', 'tube']:
                return (3.14 * ((row['BLANK WIDTH (W)'] * row['THK MM']) - (row['THK MM'] ** 2)) * row['SHEET LENGTH MM'] * 7.85 * 1e-6) / row['NO OF COMPONENTS']
            elif commodity == 'bar':
                return (3.14 * (row['BLANK WIDTH (W)'])**2 * row['SHEET LENGTH MM'] * 7.85 * 1e-6) / (4 * row['NO OF COMPONENTS'])
            else:
                return np.nan

        except (ZeroDivisionError, ValueError, TypeError):
            return np.nan
        
    
    @staticmethod
    def calculate_input_weight_actual(row):
        """Calculate actual input weight based on commodity type"""
        commodity = str(row['Commodity']).strip().lower()

        no_of_components = row.get('NO OF COMPONENTS.1', np.nan)
        if pd.isna(no_of_components) or no_of_components == 0:
            return np.nan

        try:    
            if commodity == 'plate':
                return (row['THK MM.1'] * row['SHEET WIDTH MM.1'] * row['SHEET LENGTH MM.1'] * 7.85 * 1e-6) / row['NO OF COMPONENTS.1']
            elif commodity in ['pipe', 'tube']:
                return (3.14 * ((row['BLANK WIDTH (W).1'] * row['THK MM.1']) - (row['THK MM.1'] ** 2)) * row['SHEET LENGTH MM.1'] * 7.85 * 1e-6) / row['NO OF COMPONENTS.1']
            elif commodity == 'bar':
                return (3.14 * (row['BLANK WIDTH (W).1'])**2 * row['SHEET LENGTH MM.1'] * 7.85 * 1e-6) / (4 * row['NO OF COMPONENTS.1'])
            else:
                return np.nan
        except (ZeroDivisionError, ValueError, TypeError, KeyError):
            return np.nan


# =============================================
# DATA PROCESSOR CLASS
# =============================================
class DataProcessor:
    """Handles data preprocessing and calculations"""
    
    def __init__(self, df, is_4_level=False):
        self.df = df
        self.is_4_level = is_4_level
        self.calculator = WeightCalculator()
    
    def preprocess(self):
        """Preprocess and clean data"""
        self.df = self.df.dropna()
        
        part_no_columns = [col for col in self.df.columns if 'Part No' in col]
        for col in part_no_columns:
            self.df[col] = self.df[col].astype(str).str.strip()
            self.df[col] = self.df[col].replace(['nan', 'None', ''], np.nan)

        validity_columns = [col for col in self.df.columns if 'Validity' in col]
        for col in validity_columns:
            if col in self.df.columns:
                self.df[col] = pd.to_numeric(self.df[col], errors='coerce')
                # Fill NaN validity with 1
                self.df[col] = self.df[col].fillna(1)

        columns_to_convert = [
                'Assembly Validity', 'Component Validity',
                'THK MM', 'BLANK WIDTH (W)', 'BLANK LENGTH (L)', 'NO OF COMPONENTS/BLANK',
                'SHEET WIDTH MM', 'SHEET LENGTH MM', 'NO OF COMPONENTS',
                'I/P WT. KG', 'I/P WEIGHT / CAR', 'FINISH WEIGHT IN KG',  # ← ADD THESE
                'THK MM.1', 'BLANK WIDTH (W).1', 'BLANK LENGTH (L).1', 'NO OF COMPONENTS/BLANK.1',
                'SHEET WIDTH MM.1', 'SHEET LENGTH MM.1', 'NO OF COMPONENTS.1',
                'I/P WT. KG.1', 'I/P WEIGHT / CAR.1', 'FINISH WEIGHT IN KG.1'  # ← ADD THESE
            ]
        
        if self.is_4_level:
            columns_to_convert = ['SA Validity', 'Module Validity'] + columns_to_convert
        
        for col in columns_to_convert:
            if col in self.df.columns:
                self.df[col] = self.df[col].replace([
                        '-', '--', '---', 
                        'N/A', 'n/a', 'NA', 'na',
                        'TBD', 'tbd', 'TBC', 'tbc',
                        '#DIV/0!', '#VALUE!', '#REF!', '#N/A',
                        '', ' ', '  ',
                        'None', 'none', 'null'
                    ], np.nan)

                self.df[col] = pd.to_numeric(self.df[col], errors='coerce')
                self.df[col] = self.df[col].replace([np.inf, -np.inf], np.nan)
        
        critical_actual_cols = ['THK MM.1', 'SHEET WIDTH MM.1', 'SHEET LENGTH MM.1', 'NO OF COMPONENTS.1']
        existing_critical = [col for col in critical_actual_cols if col in self.df.columns]
        
        if existing_critical:
            rows_before = len(self.df)
            
            for col in existing_critical:
                self.df = self.df[(self.df[col].notna()) & (self.df[col] != 0)]
            
            rows_after = len(self.df)
            rows_removed = rows_before - rows_after
            
            if rows_removed > 0:
                print(f"\n🗑️  Filtered out {rows_removed} rows with incomplete actual data")
                print(f"   ✅ Remaining rows: {rows_after}")
        
        # division
        division_critical_cols = [
            'NO OF COMPONENTS', 'NO OF COMPONENTS.1',
            'NO OF COMPONENTS/BLANK', 'NO OF COMPONENTS/BLANK.1'
        ]
        
        for col in division_critical_cols:
            if col in self.df.columns:
                self.df.loc[self.df[col] == 0, col] = np.nan
                self.df.loc[self.df[col] < 0, col] = np.nan
    
        calc_cols = [col for col in columns_to_convert if col in self.df.columns]
        self.df[calc_cols] = self.df[calc_cols].fillna(0)

        return self.df
    
    def perform_calculations(self):
        """Perform all weight and cost calculations"""
        has_sa = 'SA Part No' in self.df.columns

        if self.is_4_level:
            base_columns = [
                'SA Part No', 'SA Name', 'SA Validity',
                'Module Part No', 'Module Name', 'Module Validity',
                'Assembly Part No', 'Assembly Name', 'Assembly Validity',
                'Component Part No', 'Component Name', 'Component Validity',
            ]
        elif has_sa:
            # 3-level hierarchy
            base_columns = [
                'SA Part No', 'SA Name', 'SA Validity',
                'Assembly Part No', 'Assembly Name', 'Assembly Validity',
                'Component Part No', 'Component Name', 'Component Validity',
            ]
        else:
            base_columns = [
                'Assembly Part No', 'Assembly Name', 'Assembly Validity',
                'Component Part No', 'Component Name', 'Component Validity',
            ]
        
        optional_columns = ['Commodity', 'MATERIAL GRADE ']
        for col in optional_columns:
            if col in self.df.columns:
                base_columns.append(col)
        
        # Calculate validity multiplier
        if self.is_4_level:
            validity_multiplier = (self.df['SA Validity'] * self.df['Module Validity'] * 
                                self.df['Assembly Validity'] * self.df['Component Validity'])
        elif has_sa:
            validity_multiplier = (self.df['SA Validity'] * self.df['Assembly Validity'] * 
                                self.df['Component Validity'])
        else:
            validity_multiplier = self.df['Assembly Validity'] * self.df['Component Validity']
        
        df_base = self.df[base_columns].copy()
        
        # Sharada differences
        df_base['Sharada_Thickness_Difference_Actual_Theoretical'] = self.df['THK MM.1'] - self.df['THK MM']
        df_base['Sharada_Input_Weight_per_Car_Difference_Actual_Theoretical'] = (
            self.df['I/P WEIGHT / CAR.1'] - self.df['I/P WEIGHT / CAR']
        )
        df_base['Sharada_Input_weight_per_Car_Percent_Diff_Actual_Theoretical'] = (
            df_base['Sharada_Input_Weight_per_Car_Difference_Actual_Theoretical'] / self.df['I/P WEIGHT / CAR']
        ) * 100
        
        # Algo calculations
        df_base['Algo_Input_Weight_Theoretical'] = self.df.apply(
            self.calculator.calculate_input_weight_theoretical, axis=1
        )
        df_base['Algo_Input_Weight_Actual'] = self.df.apply(
            self.calculator.calculate_input_weight_actual, axis=1
        )
        
        df_base['Algo_Input_Weight_Difference_Actual_Theoretical'] = (
            df_base['Algo_Input_Weight_Actual'] - df_base['Algo_Input_Weight_Theoretical']
        )
        df_base['Algo_Input_weight_Percent_Diff_Actual_Theoretical'] = (
            df_base['Algo_Input_Weight_Difference_Actual_Theoretical'] / 
            ((df_base['Algo_Input_Weight_Actual'] + df_base['Algo_Input_Weight_Theoretical']) / 2)
        ) * 100
        
        # Per car calculations
        df_base['Algo_Input_Weight_per_CAR_Theoretical'] = (
            df_base['Algo_Input_Weight_Theoretical'] * validity_multiplier
        )
        df_base['Algo_Input_Weight_per_CAR_Actual'] = (
            df_base['Algo_Input_Weight_Actual'] * validity_multiplier
        )
        
        df_base['Algo_Input_Weight_per_CAR_Difference_Actual_Theoretical'] = (
            df_base['Algo_Input_Weight_per_CAR_Actual'] - df_base['Algo_Input_Weight_per_CAR_Theoretical']
        )
        df_base['Algo_Input_weight_per_CAR_Percent_Diff_Actual_Theoretical'] = (
            df_base['Algo_Input_Weight_per_CAR_Difference_Actual_Theoretical'] / 
            df_base['Algo_Input_Weight_per_CAR_Theoretical']
        ) * 100
        
        # Finish weight
        df_base['Algo_Finish_Weight_per_CAR'] = self.df['FINISH WEIGHT IN KG'] * validity_multiplier
        
        # Algo vs Sharada differences
        df_base['Algo_Sharada_Input_Weight_per_CAR_Difference_Theoretical'] = (
            df_base['Algo_Input_Weight_per_CAR_Theoretical'] - self.df['I/P WEIGHT / CAR']
        )
        df_base['Algo_Sharada_Input_Weight_per_CAR_Difference_Actual'] = (
            df_base['Algo_Input_Weight_per_CAR_Actual'] - self.df['I/P WEIGHT / CAR.1']
        )
        
        df_base['Algo_Sharada_Input_weight_per_CAR_Percent_Diff_Theoretical'] = (
            df_base['Algo_Sharada_Input_Weight_per_CAR_Difference_Theoretical'] / 
            ((df_base['Algo_Input_Weight_per_CAR_Theoretical'] + self.df['I/P WEIGHT / CAR']) / 2)
        ) * 100
        df_base['Algo_Sharada_Input_weight_per_CAR_Percent_Diff_Actual'] = (
            df_base['Algo_Sharada_Input_Weight_per_CAR_Difference_Actual'] / 
            ((df_base['Algo_Input_Weight_per_CAR_Actual'] + self.df['I/P WEIGHT / CAR.1']) / 2)
        ) * 100
        
        # Scrap calculations
        df_base['Algo_Scrap_Theoretical'] = (
            df_base['Algo_Input_Weight_per_CAR_Theoretical'] - df_base['Algo_Finish_Weight_per_CAR']
        )
        df_base['Algo_Scrap_Actual'] = (
            df_base['Algo_Input_Weight_per_CAR_Actual'] - df_base['Algo_Finish_Weight_per_CAR']
        )
        
        df_base['Algo_Scrap_Percent_Theoretical'] = (
            df_base['Algo_Scrap_Theoretical'] / df_base['Algo_Input_Weight_per_CAR_Theoretical']
        ) * 100
        df_base['Algo_Scrap_Percent_Actual'] = (
            df_base['Algo_Scrap_Actual'] / df_base['Algo_Input_Weight_per_CAR_Actual']
        ) * 100
        
        df_base['Algo_Scrap_Difference_Actual_Theoretical'] = (
            df_base['Algo_Scrap_Actual'] - df_base['Algo_Scrap_Theoretical']
        )
        df_base['Algo_Scrap_Percent_Difference_Actual_Theoretical'] = (
            df_base['Algo_Scrap_Difference_Actual_Theoretical'] / df_base['Algo_Scrap_Theoretical']
        ) * 100
        
        # Yield calculations
        df_base['Algo_Yield_Theoretical'] = (
            df_base['Algo_Finish_Weight_per_CAR'] / df_base['Algo_Input_Weight_per_CAR_Theoretical']
        )
        df_base['Algo_Yield_Actual'] = (
            df_base['Algo_Finish_Weight_per_CAR'] / df_base['Algo_Input_Weight_per_CAR_Actual']
        )
        
        df_base['Algo_Yield_Percent_Theoretical'] = df_base['Algo_Yield_Theoretical'] * 100
        df_base['Algo_Yield_Percent_Actual'] = df_base['Algo_Yield_Actual'] * 100
        
        df_base['Algo_Yield_Difference_Actual_Theoretical'] = (
            df_base['Algo_Yield_Actual'] - df_base['Algo_Yield_Theoretical']
        )
        df_base['Algo_Yield_Percent_Difference_Actual_Theoretical'] = (
            df_base['Algo_Yield_Difference_Actual_Theoretical'] / df_base['Algo_Yield_Theoretical']
        ) * 100
        
        return df_base


# =============================================
# MATRIX GENERATOR CLASS
# =============================================
class MatrixGenerator:
    """Handles matrix sheet creation and population"""
    
    def __init__(self, workbook, df_calculated, is_4_level=False):
        self.wb = workbook
        self.df_cal = df_calculated
        self.is_4_level = is_4_level
        self.material_grades = []
    
    def create_matrix_sheets(self):
        """Create Theoretical and Actual Matrix sheets"""
        selected_columns = ['Component Part No', 'Component Name']
        df_selected = self.df_cal[selected_columns]
        
        # Remove existing sheets if present
        for sheet_name in ['Theoretical_Matrix', 'Actual_Matrix']:
            if sheet_name in self.wb.sheetnames:
                del self.wb[sheet_name]
        
        ws_theo = self.wb.create_sheet('Theoretical_Matrix')
        ws_actual = self.wb.create_sheet('Actual_Matrix')
        
        # Write headers and data
        for col_idx, col_name in enumerate(df_selected.columns, 1):
            ws_theo.cell(1, col_idx, col_name)
            ws_actual.cell(1, col_idx, col_name)
        
        for row_idx, row in enumerate(df_selected.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws_theo.cell(row_idx, col_idx, value)
                ws_actual.cell(row_idx, col_idx, value)
        
        # Add material grades as columns
        self.material_grades = self.df_cal['MATERIAL GRADE '].dropna().unique().tolist()
        for i, grade in enumerate(self.material_grades):
            ws_theo.cell(1, 3 + i, grade)
            ws_actual.cell(1, 3 + i, grade)
        
        return self.wb
    
    def fill_matrix_data(self):
        """Fill matrix sheets with weight data"""
        ws_theo = self.wb['Theoretical_Matrix']
        ws_actual = self.wb['Actual_Matrix']
        
        # Build lookup dictionaries
        lookup_theo = defaultdict(list)
        lookup_actual = defaultdict(list)
        
        for _, row in self.df_cal.iterrows():
            part_no = row['Component Part No']
            part_name = row['Component Name']
            grade = row['MATERIAL GRADE ']
            theo_weight = row['Algo_Input_Weight_per_CAR_Theoretical']
            actual_weight = row['Algo_Input_Weight_per_CAR_Actual']
            
            lookup_theo[(part_no, part_name, grade)].append(theo_weight)
            lookup_actual[(part_no, part_name, grade)].append(actual_weight)
        
        # Fill theoretical matrix
        self._fill_sheet(ws_theo, lookup_theo)
        
        # Fill actual matrix
        self._fill_sheet(ws_actual, lookup_actual)
        
        return self.wb
    
    def _fill_sheet(self, ws, lookup):
        """Helper method to fill a worksheet with lookup data"""
        occurrence_tracker = defaultdict(int)
        row_idx = 2
        
        while ws.cell(row_idx, 1).value:
            part_no = ws.cell(row_idx, 1).value
            part_name = ws.cell(row_idx, 2).value
            
            for i, grade in enumerate(self.material_grades):
                key = (part_no, part_name, grade)
                index = occurrence_tracker[key]
                values = lookup.get(key, [])
                value = values[index] if index < len(values) else 0
                ws.cell(row_idx, 3 + i, value)
            
            for grade in self.material_grades:
                occurrence_tracker[(part_no, part_name, grade)] += 1
            row_idx += 1
    
    def add_totals(self):
        """Add total rows to matrix sheets"""
        ws_theo = self.wb['Theoretical_Matrix']
        ws_actual = self.wb['Actual_Matrix']
        
        row_idx = 2
        while ws_theo.cell(row_idx, 1).value:
            row_idx += 1
        total_row_theo = row_idx
        
        row_idx = 2
        while ws_actual.cell(row_idx, 1).value:
            row_idx += 1
        total_row_actual = row_idx
        
        ws_theo.cell(total_row_theo, 2, 'Theoretical_Total')
        ws_actual.cell(total_row_actual, 2, 'Actual_Total')
        
        for i, _ in enumerate(self.material_grades):
            col_idx = 3 + i
            
            col_sum = sum(
                ws_theo.cell(r, col_idx).value 
                for r in range(2, total_row_theo) 
                if isinstance(ws_theo.cell(r, col_idx).value, (int, float))
            )
            ws_theo.cell(total_row_theo, col_idx, col_sum)
            
            col_sum = sum(
                ws_actual.cell(r, col_idx).value 
                for r in range(2, total_row_actual) 
                if isinstance(ws_actual.cell(r, col_idx).value, (int, float))
            )
            ws_actual.cell(total_row_actual, col_idx, col_sum)
        
        return self.wb


# =============================================
# COMPARISON ANALYZER CLASS
# =============================================
class ComparisonAnalyzer:
    """Handles comparison analysis between theoretical and actual data"""
    
    def __init__(self, workbook):
        self.wb = workbook
    
    def create_comparison_sheet(self):
        """Create comparison sheet from matrix totals"""
        ws_theo = self.wb['Theoretical_Matrix']
        ws_actual = self.wb['Actual_Matrix']
        
        if 'Comparison_sheet' in self.wb.sheetnames:
            del self.wb['Comparison_sheet']
        ws_comp = self.wb.create_sheet('Comparison_sheet')
        
        material_grades = []
        col_idx = 3
        while ws_theo.cell(1, col_idx).value:
            material_grades.append(ws_theo.cell(1, col_idx).value)
            col_idx += 1
        
        theo_total_row = self._find_total_row(ws_theo, 'Theoretical_Total')
        actual_total_row = self._find_total_row(ws_actual, 'Actual_Total')
        
        headers = ['Material Grade', 'Theoretical_Total', 'Actual_Total', 'Difference', 'Difference_Percent']
        for col_idx, header in enumerate(headers, 1):
            ws_comp.cell(1, col_idx, header)
        
        for i, grade in enumerate(material_grades):
            row = i + 2
            ws_comp.cell(row, 1, grade)
            
            theo_val = ws_theo.cell(theo_total_row, 3 + i).value if theo_total_row else 0
            actual_val = ws_actual.cell(actual_total_row, 3 + i).value if actual_total_row else 0
            
            ws_comp.cell(row, 2, theo_val)
            ws_comp.cell(row, 3, actual_val)
            ws_comp.cell(row, 4, actual_val - theo_val)
            if theo_val != 0:
                ws_comp.cell(row, 5, ((actual_val - theo_val) / theo_val) * 100)
            else:
                ws_comp.cell(row, 5, 0)
        
        self._add_summary_row(ws_comp, len(material_grades))
        
        return self.wb
    
    def _find_total_row(self, ws, label):
        """Find row containing the specified label"""
        row_idx = 2
        while ws.cell(row_idx, 2).value:
            if ws.cell(row_idx, 2).value == label:
                return row_idx
            row_idx += 1
        return None
    
    def _add_summary_row(self, ws, num_grades):
        """Add summary calculations row"""
        last_row = num_grades + 2
        ws.cell(last_row, 1, 'Calculations')
        
        theo_sum = sum(ws.cell(i + 2, 2).value or 0 for i in range(num_grades))
        actual_sum = sum(ws.cell(i + 2, 3).value or 0 for i in range(num_grades))
        diff = actual_sum - theo_sum
        diff_pct = (diff / theo_sum * 100) if theo_sum != 0 else 0
        
        ws.cell(last_row, 2, theo_sum)
        ws.cell(last_row, 3, actual_sum)
        ws.cell(last_row, 4, diff)
        ws.cell(last_row, 5, diff_pct)


# =============================================
# COST ANALYZER CLASS
# =============================================
class CostAnalyzer:
    """Handles cost analysis calculations"""
    
    def __init__(self, workbook, df_full, df_calculated, steel_buy_rate, steel_scrap_rate, theoretical_margin):
        self.wb = workbook
        self.df_full = df_full
        self.df_cal = df_calculated
        self.steel_buy_rate = steel_buy_rate
        self.steel_scrap_rate = steel_scrap_rate
        self.theoretical_margin = theoretical_margin
    
    def create_cost_analysis(self):
        """Create comprehensive cost analysis"""
        ws_comp = self.wb['Comparison_sheet']
        theo_total, actual_total = self._get_totals_from_comparison(ws_comp)
        
        weight_data = self._calculate_weight_data()
        
        cost_data = self._calculate_costs(theo_total, actual_total, weight_data)
        
        result_df = self._create_result_dataframe(weight_data, cost_data)
        
        return result_df
    
    def _get_totals_from_comparison(self, ws_comp):
        """Extract totals from comparison sheet"""
        row_idx = 2
        while ws_comp.cell(row_idx, 1).value:
            if str(ws_comp.cell(row_idx, 1).value).lower() == 'calculations':
                return ws_comp.cell(row_idx, 2).value, ws_comp.cell(row_idx, 3).value
            row_idx += 1
        return None, None
    
    def _calculate_weight_data(self):
        """Calculate all weight-related metrics"""
        return {
            'sharada_theo_input': self.df_full['I/P WT. KG'].sum(),
            'algo_theo_input': self.df_cal['Algo_Input_Weight_Theoretical'].sum(),
            'sharada_actual_input': self.df_full['I/P WT. KG.1'].sum(),
            'algo_actual_input': self.df_cal['Algo_Input_Weight_Actual'].sum(),
            'sharada_theo_input_car': self.df_full['I/P WEIGHT / CAR'].sum(),
            'algo_theo_input_car': self.df_cal['Algo_Input_Weight_per_CAR_Theoretical'].sum(),
            'sharada_actual_input_car': self.df_full['I/P WEIGHT / CAR.1'].sum(),
            'algo_actual_input_car': self.df_cal['Algo_Input_Weight_per_CAR_Actual'].sum(),
            'total_finish_car': self.df_cal['Algo_Finish_Weight_per_CAR'].sum()
        }
    
    def _calculate_costs(self, theo_total, actual_total, weight_data):
        """Calculate all cost-related metrics"""
        theo_scrap = theo_total - weight_data['total_finish_car']
        actual_scrap = actual_total - weight_data['total_finish_car']
        
        theo_cost = theo_total * self.steel_buy_rate
        actual_cost = actual_total * self.steel_buy_rate
        
        theo_scrap_value = theo_scrap * self.steel_scrap_rate
        actual_scrap_value = actual_scrap * self.steel_scrap_rate
        
        additional_diff = actual_scrap - theo_scrap
        labor_cost = theo_cost * 0.2
        tata_invoice = (theo_cost + labor_cost - theo_scrap_value) * (1 + self.theoretical_margin)
        actual_cost_incurred = actual_cost + labor_cost - actual_scrap_value
        actual_margin = (tata_invoice - actual_cost_incurred) / tata_invoice if tata_invoice != 0 else 0
        
        return {
            'theo_total': theo_total,
            'actual_total': actual_total,
            'theo_scrap': theo_scrap,
            'actual_scrap': actual_scrap,
            'theo_cost': theo_cost,
            'actual_cost': actual_cost,
            'theo_scrap_value': theo_scrap_value,
            'actual_scrap_value': actual_scrap_value,
            'additional_diff': additional_diff,
            'additional_steel_value': additional_diff * self.steel_buy_rate,
            'additional_scrap_value': additional_diff * self.steel_scrap_rate,
            'labor_cost': labor_cost,
            'tata_invoice': tata_invoice,
            'actual_cost_incurred': actual_cost_incurred,
            'actual_margin': actual_margin
        }
    
    def _create_result_dataframe(self, weight_data, cost_data):
        """Create the final cost analysis dataframe"""
        
        # Helper function to clean tiny values - MORE AGGRESSIVE
        def clean_value(val, precision=6):
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                if abs(val) < 1e-6:  # Changed from 1e-9
                    return 0
                return round(val, precision)
            return val
        
        # Calculate differences (Algo - Sharada, as per original code)
        # diff_theo_input = clean_value(weight_data['algo_theo_input'] - weight_data['sharada_theo_input'])
        # diff_actual_input = clean_value(weight_data['algo_actual_input'] - weight_data['sharada_actual_input'])
        # diff_theo_input_car = clean_value(weight_data['algo_theo_input_car'] - weight_data['sharada_theo_input_car'])
        # diff_actual_input_car = clean_value(weight_data['algo_actual_input_car'] - weight_data['sharada_actual_input_car'])
        
        result_df = pd.DataFrame({
            "Metric": [
                "Total Input Weight Sharada (Kg)",
                # "Total Input Weight Algo (Kg)",
                # "Difference Input Weight (Kg)",
                "Total Input Weight Per Car Sharada (Kg)",
                # "Total Input Weight Per Car Algo (Kg)",
                # "Difference Input Weight Per Car (Kg)",
                "Steel procurement Per Car (Kg)",
                "Steel Scrap (Kg)",
                "Steel procurement Cost (INR)",
                "Scrap Value (INR)",
                "Tata Motors Invoice (INR) / Actual Cost Incurred (INR)",
                "Margin (%)",
            ],
            "Theoretical": [
                clean_value(weight_data['sharada_theo_input']),
                # clean_value(weight_data['algo_theo_input']),
                # diff_theo_input,
                clean_value(weight_data['sharada_theo_input_car']),
                # clean_value(weight_data['algo_theo_input_car']),
                # diff_theo_input_car,
                clean_value(cost_data['theo_total']),
                clean_value(cost_data['theo_scrap']),
                clean_value(cost_data['theo_cost'], 2),  # Currency: 2 decimals
                clean_value(cost_data['theo_scrap_value'], 2),
                clean_value(cost_data['tata_invoice'], 2),
                clean_value(self.theoretical_margin * 100, 2),
            ],
            "Actual": [
                clean_value(weight_data['sharada_actual_input']),
                # clean_value(weight_data['algo_actual_input']),
                # diff_actual_input,
                clean_value(weight_data['sharada_actual_input_car']),
                # clean_value(weight_data['algo_actual_input_car']),
                # diff_actual_input_car,
                clean_value(cost_data['actual_total']),
                clean_value(cost_data['actual_scrap']),
                clean_value(cost_data['actual_cost'], 2),
                clean_value(cost_data['actual_scrap_value'], 2),
                clean_value(cost_data['actual_cost_incurred'], 2),
                clean_value(cost_data['actual_margin'] * 100, 2),
            ],
            "": [""] * 8,
            "Constants/Metrics": [
                "Total Finish Weight Per Car (Kg)",
                "Steel Buy Rate (INR/kg)",
                "Steel Scrap Rate (INR/kg)",
                "Additional Difference (KG)",
                "Additional Steel Value (INR)",
                "Additional Scrap Value (INR)",
                "Overhead Cost (INR)",
                "",
            ],
            "Value": [
                clean_value(weight_data['total_finish_car']),
                clean_value(self.steel_buy_rate, 2),
                clean_value(self.steel_scrap_rate, 2),
                clean_value(cost_data['additional_diff']),
                clean_value(cost_data['additional_steel_value'], 2),
                clean_value(cost_data['additional_scrap_value'], 2),
                clean_value(cost_data['labor_cost'], 2),
                "",
            ]
        })
        
        return result_df


# =============================================
# FILTERED COMPONENTS GENERATOR CLASS
# =============================================
class FilteredComponentsGenerator:
    """Handles creation of filtered components with deviation analysis"""
    
    def __init__(self, df_full, df_calculated, is_4_level=False):
        self.df_full = df_full
        self.df_cal = df_calculated
        self.is_4_level = is_4_level
    
    def create_filtered_components(self):
        """Create filtered components sheet with highlighting"""
        # Part 1: Basic info
        part1_cols = ['Component Part No', 'Component Name', 'Component Validity', 
                      'Commodity', 'MATERIAL GRADE ']
        df_part1 = self.df_full[part1_cols].copy()
        
        # Part 2: Weight data
        df_part2 = self._create_weight_section()
        
        # Part 3: Dimensional comparisons
        df_part3 = self._create_dimensional_section()
        
        # Combine all parts
        result_df = pd.concat([df_part1, df_part2, df_part3], axis=1)
        result_df = result_df.sort_values(by='% Deviation (Input Weight Per Car)', key=abs, ascending=False)
        
        return result_df
    
    def _create_weight_section(self):
        """Create weight comparison section"""
        df_part2 = pd.DataFrame({
            'Input Weight Per Car (Theoretical)': self.df_cal['Algo_Input_Weight_per_CAR_Theoretical'],
            'Input Weight Per Car (Actual)': self.df_cal['Algo_Input_Weight_per_CAR_Actual']
        })
        
        difference = df_part2['Input Weight Per Car (Actual)'] - df_part2['Input Weight Per Car (Theoretical)']
        pct_diff = (difference / df_part2['Input Weight Per Car (Theoretical)']) * 100
        
        df_part2['Difference_Input_Weight_Per_Car'] = difference
        df_part2['% Deviation (Input Weight Per Car)'] = pct_diff
        
        return df_part2
    
    def _create_dimensional_section(self):
        """Create dimensional comparison section"""
        dim_cols = [
            ('THK MM', 'THK MM.1', 'THK (Theoretical)', 'THK (Actual)', 'Difference_THK', '% Deviation_THK'),
            ('BLANK WIDTH (W)', 'BLANK WIDTH (W).1', 'BLANK WIDTH (Theoretical)', 'BLANK WIDTH (Actual)', 
             'Difference_BLANK_WIDTH', '% Deviation_BLANK_WIDTH'),
            ('BLANK LENGTH (L)', 'BLANK LENGTH (L).1', 'BLANK LENGTH (Theoretical)', 'BLANK LENGTH (Actual)', 
             'Difference_BLANK_LENGTH', '% Deviation_BLANK_LENGTH'),
            ('NO OF COMPONENTS/BLANK', 'NO OF COMPONENTS/BLANK.1', 'NO OF COMPONENTS/BLANK (Theoretical)', 
             'NO OF COMPONENTS/BLANK (Actual)', 'Difference_NO_OF_COMPONENTS/BLANK', '% Deviation_NO_OF_COMPONENTS/BLANK'),
            ('SHEET WIDTH MM', 'SHEET WIDTH MM.1', 'SHEET WIDTH MM (Theoretical)', 'SHEET WIDTH MM (Actual)', 
             'Difference_SHEET_WIDTH', '% Deviation_SHEET_WIDTH'),
            ('SHEET LENGTH MM', 'SHEET LENGTH MM.1', 'SHEET LENGTH MM (Theoretical)', 'SHEET LENGTH MM (Actual)', 
             'Difference_SHEET_LENGTH', '% Deviation_SHEET_LENGTH'),
            ('NO OF COMPONENTS', 'NO OF COMPONENTS.1', 'NO OF COMPONENTS (Theoretical)', 'NO OF COMPONENTS (Actual)', 
             'Difference_NO_OF_COMPONENTS', '% Deviation_NO_OF_COMPONENTS'),
        ]
        
        df_part3 = pd.DataFrame()
        for col_theo, col_act, name_theo, name_act, name_diff, name_pct in dim_cols:
            df_part3[name_theo] = self.df_full[col_theo]
            df_part3[name_act] = self.df_full[col_act]
            df_part3[name_diff] = self.df_full[col_act] - self.df_full[col_theo]
            with pd.option_context('mode.use_inf_as_na', True):
                df_part3[name_pct] = ((self.df_full[col_act] - self.df_full[col_theo]) / self.df_full[col_theo]) * 100
        
        return df_part3


# =============================================
# EXCEL STYLER CLASS
# =============================================
class ExcelStyler:
    """Handles Excel cell highlighting and formatting"""
    
    @staticmethod
    def apply_styles_to_workbook(workbook, df_calculated):
        """Apply conditional formatting styles to the Calculated_Data sheet"""
        from openpyxl.styles import PatternFill
        
        # Get the Calculated_Data sheet
        if 'Calculated_Data' not in workbook.sheetnames:
            return workbook
        
        ws = workbook['Calculated_Data']
        
        # Define fills
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        lightcoral_fill = PatternFill(start_color='F08080', end_color='F08080', fill_type='solid')
        
        # Get column indices (create a mapping of column names to indices)
        header_row = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
        
        # Define highlighting rules based on your original functions
        highlighting_rules = {
            # Rule 1: highlight_sharada_thickness_differences
            'Sharada_Thickness_Difference_Actual_Theoretical': lambda val: (
                yellow_fill if round(val, 6) < -0.05 else
                orange_fill if round(val, 6) > 0.05 else None
            ),
            
            # Rule 2: highlight_sharada_algo_input_percent_differences
            'Sharada_Input_weight_per_Car_Percent_Diff_Actual_Theoretical': lambda val: (
                yellow_fill if round(val, 6) <= -5 else
                orange_fill if round(val, 6) >= 5 else None
            ),
            'Algo_Input_weight_per_CAR_Percent_Diff_Actual_Theoretical': lambda val: (
                yellow_fill if round(val, 6) <= -5 else
                orange_fill if round(val, 6) >= 5 else None
            ),
            
            # Rule 3: highlight_percent_input_differences
            'Algo_Sharada_Input_weight_per_CAR_Percent_Diff_Theoretical': lambda val: (
                yellow_fill if round(val, 6) <= -5 else
                orange_fill if round(val, 6) >= 5 else None
            ),
            'Algo_Sharada_Input_weight_per_CAR_Percent_Diff_Actual': lambda val: (
                yellow_fill if round(val, 6) <= -5 else
                orange_fill if round(val, 6) >= 5 else None
            ),
            
            # Rule 4: highlight_negative_values (for Scrap Percent)
            'Algo_Scrap_Percent_Theoretical': lambda val: (
                yellow_fill if val < 0 else
                lightcoral_fill if val > 50 else None
            ),
            'Algo_Scrap_Percent_Actual': lambda val: (
                yellow_fill if val < 0 else
                lightcoral_fill if val > 50 else None
            ),
            
            # Rule 5: highlight_negative_yeild_values (for Yield Percent)
            'Algo_Yield_Percent_Theoretical': lambda val: (
                yellow_fill if val < 0 else
                lightcoral_fill if val < 50 else None
            ),
            'Algo_Yield_Percent_Actual': lambda val: (
                yellow_fill if val < 0 else
                lightcoral_fill if val < 50 else None
            ),
            
            # Rule 6: highlight_percent_scrap_yield_differences
            'Algo_Scrap_Percent_Difference_Actual_Theoretical': lambda val: (
                yellow_fill if round(val, 6) <= -5 else
                orange_fill if round(val, 6) >= 5 else None
            ),
            'Algo_Yield_Percent_Difference_Actual_Theoretical': lambda val: (
                yellow_fill if round(val, 6) <= -5 else
                orange_fill if round(val, 6) >= 5 else None
            ),
        }
        
        # Apply highlighting to each cell
        for row_idx in range(2, ws.max_row + 1):
            for col_name, highlight_func in highlighting_rules.items():
                if col_name in header_row:
                    col_idx = header_row[col_name]
                    cell = ws.cell(row_idx, col_idx)
                    
                    # Apply highlighting if cell has a numeric value
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        try:
                            fill = highlight_func(cell.value)
                            if fill is not None:
                                cell.fill = fill
                        except Exception:
                            # Skip cells that cause errors (e.g., NaN, inf)
                            pass


    
        
        return workbook
    
    @staticmethod
    def apply_filtered_components_styling(workbook):
        """Apply conditional formatting to Filtered_Components sheet"""
        from openpyxl.styles import PatternFill
        
        if 'Filtered_Components' not in workbook.sheetnames:
            return workbook
        
        ws = workbook['Filtered_Components']
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        
        header_row = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
        
        deviation_columns = [
            '% Deviation_THK',
            '% Deviation_BLANK_WIDTH',
            '% Deviation_BLANK_LENGTH',
            '% Deviation_NO_OF_COMPONENTS/BLANK',
            '% Deviation_SHEET_WIDTH',
            '% Deviation_SHEET_LENGTH',
            '% Deviation_NO_OF_COMPONENTS'
        ]
        
        highlight_column = '% Deviation (Input Weight Per Car)'
        
        dev_col_indices = {}
        for col in deviation_columns:
            if col in header_row:
                dev_col_indices[col] = header_row[col]
        
        highlight_col_idx = header_row.get(highlight_column)
        
        for row_idx in range(2, ws.max_row + 1):
            if highlight_col_idx:
                cell = ws.cell(row_idx, highlight_col_idx)
                if cell.value is not None:
                    cell.fill = orange_fill
            
            max_val = None
            max_col_idx = None
            
            for col_name, col_idx in dev_col_indices.items():
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    abs_val = abs(cell.value)
                    if max_val is None or abs_val > max_val:
                        max_val = abs_val
                        max_col_idx = col_idx
            
            if max_col_idx is not None:
                ws.cell(row_idx, max_col_idx).fill = yellow_fill
        
        return workbook


    @staticmethod
    def apply_cost_analysis_styling(workbook):
        """Apply conditional formatting to Cost_Analysis sheet"""
        from openpyxl.styles import PatternFill
        
        if 'Cost_Analysis' not in workbook.sheetnames:
            return workbook
        
        ws = workbook['Cost_Analysis']
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        margin_row_idx = None
        actual_col_idx = None
        
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value and str(cell_value).strip() == "Margin (%)":
                margin_row_idx = row_idx
                break
        
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(1, col_idx).value
            if cell_value and str(cell_value).strip() == "Actual":
                actual_col_idx = col_idx
                break
        
        if margin_row_idx and actual_col_idx:
            cell = ws.cell(margin_row_idx, actual_col_idx)
            cell.fill = yellow_fill
        
        return workbook

# =============================================
# PROCESSOR CLASS
# =============================================
class Pipeline:
    """Main orchestrator class for the entire processing pipeline"""
    
    def __init__(self, uploaded_file, sheet_index, is_4_level, steel_buy_rate, steel_scrap_rate, theoretical_margin):
        self.uploaded_file = uploaded_file
        self.sheet_index = sheet_index
        self.is_4_level = is_4_level
        self.steel_buy_rate = steel_buy_rate
        self.steel_scrap_rate = steel_scrap_rate
        self.theoretical_margin = theoretical_margin
        
        self.df_input = None
        self.df_flat = None
        self.df_calculated = None
        self.wb = None
    
    @staticmethod
    @st.cache_data(show_spinner=False, max_entries=5)
    def _cached_read_excel(file_bytes, sheet_index):
        return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_index)
    

    @staticmethod
    def normalize_column_names(df):
        """Normalize column names to handle variations"""
        column_mapping = {}
        
        for col in df.columns:
            original = col
            col_str = str(col).strip()
            col_lower = col_str.lower()
            
            suffix = ''
            base_name = col_str
            if '.' in col_str and col_str.split('.')[-1].isdigit():
                parts = col_str.rsplit('.', 1)
                base_name = parts[0]
                suffix = '.' + parts[1]
            
            base_lower = base_name.lower()
            
            if 'material' in base_lower and 'grade' in base_lower:
                normalized = 'MATERIAL GRADE ' + suffix
            
            elif 'commod' in base_lower:
                normalized = 'Commodity' + suffix
            
            else:
                variations = {
                    'part no': 'PART NO',
                    'partno': 'PART NO',
                    'part number': 'PART NO',
                    'part name': 'PART NAME',
                    'partname': 'PART NAME',
                    'type': 'Type',
                    'mod': 'MOD',
                    'comodity':'Commodity',
                    'validity': 'Validity',
                    'level': 'LEVEL',
                    
                    # Dimensions
                    'thk mm': 'THK MM',
                    'thkmm': 'THK MM',
                    'thickness mm': 'THK MM',
                    'thickness': 'THK MM',
                    
                    'blank width': 'BLANK WIDTH (W)',
                    'blank width (w)': 'BLANK WIDTH (W)',
                    'blankwidth': 'BLANK WIDTH (W)',
                    
                    'blank length': 'BLANK LENGTH (L)',
                    'blank length (l)': 'BLANK LENGTH (L)',
                    'blanklength': 'BLANK LENGTH (L)',
                    
                    'no of components/blank': 'NO OF COMPONENTS/BLANK',
                    'noofcomponents/blank': 'NO OF COMPONENTS/BLANK',
                    
                    'sheet width mm': 'SHEET WIDTH MM',
                    'sheetwidthmm': 'SHEET WIDTH MM',
                    'sheet width': 'SHEET WIDTH MM',
                    
                    'sheet length mm': 'SHEET LENGTH MM',
                    'sheetlengthmm': 'SHEET LENGTH MM',
                    'sheet length': 'SHEET LENGTH MM',
                    
                    'no of components': 'NO OF COMPONENTS',
                    'noofcomponents': 'NO OF COMPONENTS',
                    
                    # Weights
                    'i/p wt. kg': 'I/P WT. KG',
                    'i/p wt kg': 'I/P WT. KG',
                    'input weight kg': 'I/P WT. KG',
                    
                    'i/p weight / car': 'I/P WEIGHT / CAR',
                    'i/p weight/car': 'I/P WEIGHT / CAR',
                    'input weight / car': 'I/P WEIGHT / CAR',
                    
                    'finish weight in kg': 'FINISH WEIGHT IN KG',
                    'finishweightinkg': 'FINISH WEIGHT IN KG',
                    'finish weight kg': 'FINISH WEIGHT IN KG',
                    
                    'fin weight / car': 'FIN WEIGHT / CAR',
                    'fin weight/car': 'FIN WEIGHT / CAR',
                    
                    # Other
                    'yield': 'YIELD',
                    'scrap recovery': 'SCRAP RECOVERY',
                    'scraprecovery': 'SCRAP RECOVERY',
                }
                
                if base_lower in variations:
                    normalized = variations[base_lower] + suffix
                else:
                    normalized = original  
            
            column_mapping[original] = normalized
        
        df = df.rename(columns=column_mapping)
        
        return df

    def get_hierarchy_info(self):
        """Get detected hierarchy information"""
        has_sa = 'SA Part No' in self.df_flat.columns if self.df_flat is not None else False
        has_module = 'Module Part No' in self.df_flat.columns if self.df_flat is not None else False
        
        if self.is_4_level and has_module:
            return "4 Levels (SA → Module → Assembly → Component)"
        elif has_sa:
            return "3 Levels (SA → Assembly → Component)"
        else:
            return "2 Levels (Assembly → Component)"

    def process(self):
        """Execute the complete processing pipeline"""
        file_bytes = self.uploaded_file.read()
        self.uploaded_file.seek(0)  # Reset file pointer
        self.df_input = self._cached_read_excel(file_bytes, self.sheet_index)
        
        self.df_input = self.normalize_column_names(self.df_input)

        flattener = DataFlattener(self.df_input)
        self.df_flat, detected_is_4_level = flattener.flatten_auto()
        self.is_4_level = detected_is_4_level
        
        processor = DataProcessor(self.df_flat, self.is_4_level)
        self.df_flat = processor.preprocess()
        self.df_calculated = processor.perform_calculations()
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            self.df_flat.to_excel(writer, sheet_name='Full Data', index=False)
            self.df_calculated.to_excel(writer, sheet_name='Calculated_Data', index=False)
        
        output.seek(0)
        self.wb = load_workbook(output)
        
        matrix_gen = MatrixGenerator(self.wb, self.df_calculated, self.is_4_level)
        self.wb = matrix_gen.create_matrix_sheets()
        self.wb = matrix_gen.fill_matrix_data()
        self.wb = matrix_gen.add_totals()
        
        comparison = ComparisonAnalyzer(self.wb)
        self.wb = comparison.create_comparison_sheet()
        
        styler = ExcelStyler()
        self.wb = styler.apply_styles_to_workbook(self.wb, self.df_calculated)
        
        cost_analyzer = CostAnalyzer(
            self.wb, self.df_flat, self.df_calculated,
            self.steel_buy_rate, self.steel_scrap_rate, self.theoretical_margin
        )
        cost_df = cost_analyzer.create_cost_analysis()
        
        self._add_sheet_from_dataframe('Cost_Analysis', cost_df)

        styler = ExcelStyler()
        self.wb = styler.apply_cost_analysis_styling(self.wb)
        
        filtered_gen = FilteredComponentsGenerator(self.df_flat, self.df_calculated, self.is_4_level)
        filtered_df = filtered_gen.create_filtered_components()
        
        self._add_sheet_from_dataframe('Filtered_Components', filtered_df)
        
        self.wb = styler.apply_filtered_components_styling(self.wb)

        final_output = io.BytesIO()
        self.wb.save(final_output)
        final_output.seek(0)
        
        del self.wb
        return final_output, cost_df, filtered_df
    
    def _add_sheet_from_dataframe(self, sheet_name, df):
        """Helper method to add a dataframe as a sheet to the workbook"""
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]
        
        ws = self.wb.create_sheet(sheet_name)
        
        for c_idx, col_name in enumerate(df.columns.tolist(), 1):
            ws.cell(1, c_idx, col_name)
        
        for r_idx, row in enumerate(df.values.tolist(), 2):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    if abs(value) < 1e-6:  
                        ws.cell(r_idx, c_idx, 0)
                    else:
                        ws.cell(r_idx, c_idx, round(value, 6))
                else:
                    ws.cell(r_idx, c_idx, value)
    
    def get_metrics(self):
        """Get key metrics for display"""
        return {
            'total_rows': len(self.df_flat),
            'total_weight': self.df_calculated['Algo_Input_Weight_per_CAR_Theoretical'].sum(),
            'material_grades': len(self.df_calculated['MATERIAL GRADE '].dropna().unique())
            # 'sheets_created': len(self.wb.sheetnames)
        }



# =============================================
# STREAMLIT APP CLASS
# =============================================
class DashboardApp:
    """Main Streamlit application class"""
    
    def __init__(self):
        self.setup_page()
    
    def setup_page(self):
        """Configure Streamlit page settings"""
        st.set_page_config(page_title="Theoretical vs Actual", layout="wide", page_icon="🏭")
        st.title("Theoretical vs Actual Calculator")
        st.markdown("---")
    
    def display_metrics(self, metrics, hierarchy_info):
        """Display key metrics"""
        st.success("✅ Processing complete!")  
        st.info(f"🔍 **Detected Hierarchy:** {hierarchy_info}")
        
    
    def display_results(self, cost_df, filtered_df, df_input):
        """Display analysis results"""
        st.markdown("---")
        st.markdown("### 📋 Analysis Results")
        
        tab1, tab2, tab3 = st.tabs(["📊 Cost Analysis", "🔍 Top Deviations", "📄 Input Preview"])
        
        with tab1:
            def highlight_actual_margin(df):
                """Highlight Actual Margin % cell"""
                styles = pd.DataFrame('', index=df.index, columns=df.columns)
                
                # Find Margin (%) row
                margin_row = df[df['Metric'] == 'Margin (%)'].index
                if len(margin_row) > 0 and 'Actual' in df.columns:
                    styles.loc[margin_row[0], 'Actual'] = 'background-color: yellow'
                
                return styles
            
            styled_cost_df = cost_df.style.apply(highlight_actual_margin, axis=None)
            st.dataframe(styled_cost_df, use_container_width=True)
        
        with tab2:
            def highlight_filtered_components(row):
                """Highlight max deviation and Input Weight Per Car deviation"""
                styles = [''] * len(row)
                
                deviation_columns = [
                    '% Deviation_THK',
                    '% Deviation_BLANK_WIDTH',
                    '% Deviation_BLANK_LENGTH',
                    '% Deviation_NO_OF_COMPONENTS/BLANK',
                    '% Deviation_SHEET_WIDTH',
                    '% Deviation_SHEET_LENGTH',
                    '% Deviation_NO_OF_COMPONENTS'
                ]
                
                highlight_column = '% Deviation (Input Weight Per Car)'
                
                if all(col not in row.index for col in deviation_columns) and highlight_column not in row.index:
                    return styles
                
                for i, col in enumerate(row.index):
                    if col == highlight_column:
                        styles[i] = 'background-color: orange'
                
                try:
                    values = row[deviation_columns].abs()
                    if not values.isnull().all():
                        max_val = values.max()
                        for i, col in enumerate(row.index):
                            if col in deviation_columns and abs(row[col]) == max_val:
                                styles[i] = 'background-color: yellow'
                except:
                    pass
                
                return styles
            
            styled_filtered_df = filtered_df.head(20).style.apply(highlight_filtered_components, axis=1)
            st.dataframe(styled_filtered_df, use_container_width=True)
        
        with tab3:
            if df_input is not None:
                st.dataframe(df_input.head(20), use_container_width=True)
            else:
                st.info("Input data preview not available")
    
    def run(self):
        """Main application execution"""
        uploaded_file = st.file_uploader(
            "📤 Upload Excel File", 
            type=['xlsx', 'xls'],
            help="Select your Excel file to process"
        )
        
        if uploaded_file is not None:
            col1, col2 = st.columns([3, 1])
            with col1:
                st.info(f"📁 **File:** {uploaded_file.name}")
            with col2:
                sheet_index = 0
            
            if st.button("🚀 Process File", type="primary", use_container_width=True):
                try:
                    with st.spinner("⏳ Processing your file..."):
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        status_text.text("🔄 Initializing processor...")
                        progress_bar.progress(10)
                        
                        pipeline = Pipeline(
                            uploaded_file,
                            sheet_index,
                            False,  
                            STEEL_BUY_RATE,
                            STEEL_SCRAP_RATE,
                            THEORETICAL_MARGIN
                        )
                        
                        status_text.text("🔄 Reading and flattening data...")
                        progress_bar.progress(30)
                        
                        status_text.text("🔄 Performing calculations...")
                        progress_bar.progress(50)
                        
                        status_text.text("🔄 Creating matrices and analysis...")
                        progress_bar.progress(70)
                        
                        final_output, cost_df, filtered_df = pipeline.process()
                        
                        status_text.text("🔄 Finalizing...")
                        progress_bar.progress(90)
                        
                        metrics = pipeline.get_metrics()
                        hierarchy_info = pipeline.get_hierarchy_info()
                        
                        progress_bar.progress(100)
                        status_text.empty()
                        progress_bar.empty()
                        
                        self.display_metrics(metrics, hierarchy_info)
                        self.display_results(cost_df, filtered_df, pipeline.df_input)
                        
                        st.markdown("---")
                        st.download_button(
                            label="📥 Download Processed Excel File",
                            data=final_output,
                            file_name=f"output.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            use_container_width=True
                        )
                        
                except Exception as e:
                    st.error(f"❌ An error occurred: {str(e)}")
                    with st.expander("🔍 View Error Details"):
                        st.exception(e)


# =============================================
# MAIN ENTRY POINT
# =============================================
if __name__ == "__main__":
    app = DashboardApp()
    app.run()